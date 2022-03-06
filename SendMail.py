from fileinput import filename
from account import MY_ID, MY_PW # account라는 py 파일을 만든 뒤 MY_ID, MY_PW 로 id와 app pw 입력
from email.message import EmailMessage
from smtplib import SMTP_SSL
from openpyxl import load_workbook
from pathlib import Path

def send_mail(ToPeople, Title, Content, AttachFile=False):
    
    mail = EmailMessage()

    # 보내는 사람 / 받는 사람 / 제목 입력
    mail["From"] = MY_ID
    mail["To"] = ToPeople
    mail["Subject"] = Title

    mail.set_content(Content)

    if AttachFile:
        fName = Path(AttachFile).name # 해당 경로의 파일 이름 가져오기
        with open(AttachFile, 'rb') as file:
            # add_attachment(파일.read(), maintype/subtype[ text/plain, iamge/jpeg, application/pdf ...], 파일 이름)
            mail.add_attachment(file.read(), maintype="application", subtype="octet-stream", filename=fName)
            mail.add_header('Content-Disposition', 'attachment', filename=fName)

    with SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(MY_ID, MY_PW)
        smtp.send_message(mail)


filepath = "엑셀 파일 경로"
excelfile = load_workbook(filepath, data_only=True)   # 엑셀 파일 로드
MList = excelfile.active # 현재 활성화된 Sheet 얻음

# 엑셀파일에서 가져온 정보를 이용해 반복 실행
for row in MList.iter_rows(min_row=2): # min_row = 읽기 시작할 행 번호 / 2번째 행부터 시작
    MtoPeople = row[0].value
    MTitle = row[1].value
    MContent = row[2].value
    MAttachFile = row[3].value
    send_mail(MtoPeople, MTitle, MContent, MAttachFile)